#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jun 28 17:54:34 2023

@author: aiden
"""
import vibfreq
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabel, DataLabelList
import matplotlib.pyplot as plt
from timeit import default_timer as timer
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.text import RichText

import math


# make sure to use:
#   pip install -r requirements.txt
# before running the first time

INPUT_FILES = {  # dictionary containing molecule names and where the input file is located
    "isoquinoline1":"../test/1-butylnaptho[2-3-g]isoquinoline.log"
}

OUTPUT_EXCEL_FILE = "test.xlsx"

# GaussSum Parameters
START = 8   # note: endpoints are included so step may not be intuitive to calculate
END = 4000
NUM_PTS = 500
FWHM = 10
def SCALE_FUNCTION(freq):
    if freq < 1111.11:
        scalingFactor = 0.979
    elif freq > 2500:
        scalingFactor = 0.961
    else:
        scalingFactor = 0.973    
    return scalingFactor
    
EXCITATION = 785
TEMP = 293.15


# Excel Parameters
DOFFSET = 100  # the amount of offset to add to each molecule


##########################
#
# Peak finding parameters
#
##########################
WINDOW_SIZE = 20     # look around
N_SIGMA = .5         # sensitivity cutoff
COALESCE_WINDOW = 9  # coalesce
HIGH_PASS = 9

# this parameter must be a reasonable value in order for the assumption that
# the wavelength series is evenly spaced to be approximately true because the
# peak finding algorithm used and most peak finding algorithms in general require
# this
WAV_X_MAX = 25
WAV_X_MIN = 2.5
FREQ_X_MAX = 4000
FREQ_X_MIN = 0


AXIS_TITLE_FONT_SIZE = 16
AXIS_FONT_SIZE = 14
LABEL_FONT_SIZE = 14


# exclusive regions of where all local maxes should be included
INCLUDE_ALL_LOCAL_MAX_RANGES = [
    (6.0, 6.4),
    (12, 13)
]


##########################
#
# Misc parameters
#
##########################
MPL_PLOT = False  # set true if you would like to view which peaks were chosen
                  # before having to open excel



# anything below this line is not to be modified by typical users

start = timer()
workingDir = "./"



###############################################################################
#
#                         GaussSum
#
###############################################################################

for moleculeName, inputFile in INPUT_FILES.items():
    outputFile = workingDir + moleculeName + ".out"
    vibfreq.irSpectra(inputFile, outputFile, START, END, NUM_PTS, FWHM, SCALE_FUNCTION)


###############################################################################
#
#                         Peak Finding
#
###############################################################################

def mean(data):
    return sum(data) / len(data)

def stdev(data):
    m = mean(data)
    s = 0
    for d in data:
        s += (d - m)**2

    return math.sqrt(s / len(data))

def getWindow(data, center, sideLength):
    if center + sideLength < len(data) and center - sideLength > 0:
        return data[center - sideLength : center + sideLength + 1]
    elif center + sideLength >= len(data):
        return data[len(data) - 1 - sideLength * 2:]
    else:
        return data[0:sideLength * 2 + 1]

def isLocalMax(data, i):
    if i > 0 and i < len(data) - 1:
        return (data[i] > data[i + 1] and data[i] > data[i - 1])
    else:
        return False

# peak algorithm based upon this research paper:
# https://www.researchgate.net/publication/228853276_Simple_Algorithms_for_Peak_Detection_in_Time-Series
# An example of its operation is shown here:
# https://observablehq.com/@yurivish/peak-detection
# Method assumes an evenly spaced time series, this is approximately true for
# the usable range of the data we are looking at
def findPeaks(dataX, dataY, max_x, windowSize, nsigma, coalesceWindow, highPass, molName, plot):
    peaks = []
    lastPeakIndex = -1
    for i in range(len(dataY)):
        if dataX[i] > max_x:  # data from highest to lowest and reversing
                               # will mess up indexing, so just skip it
            continue


        # peak was not added, so check to see if is in one of the include-all-local-max
        # regions
        xVal = dataX[i]
        if any([xVal > r[0] and xVal < r[1] for r in INCLUDE_ALL_LOCAL_MAX_RANGES]):
            if isLocalMax(dataY, i):
                peaks.append(i)
                lastPeakIndex = -1  # reset to negative number so no coalesce
                continue
            
            
        window = getWindow(dataY, i, windowSize)
        m = mean(window)
        s = stdev(window)

        # potential peak, so make sure it is not too close to another peak
        if dataY[i] - m > nsigma * s and dataY[i] > highPass:
            if lastPeakIndex > 0:  # there is a previous peak
                if i - lastPeakIndex < coalesceWindow:  # choose maximum if within window
                    if dataY[i] > dataY[lastPeakIndex]:
                        peaks[-1] = i
                        lastPeakIndex = i
                    # no changes made otherwise
                else:
                    peaks.append(i)
                    lastPeakIndex = i
            else:
                peaks.append(i)
                lastPeakIndex = i
                
    if plot:
        print([dataX[i] for i in peaks])
        fig = plt.figure()
        ax = fig.add_subplot(111)
        ax.plot(dataX, dataY)
        ax.scatter([dataX[i] for i in peaks], [dataY[i] for i in peaks], color="green")
        ax.set_xlim(0, max_x)
        fig.suptitle(molName)
        plt.pause(5)

    return peaks



###############################################################################
#
#                         Excel Data Dump
#
###############################################################################
wb = Workbook()
wb.remove(wb.active)

wb.create_sheet("config")

configRow = 1
currentOffset = 0
moleculeData = {}  # {molecule name: {dataLabel: [data]}}
for moleculeName, inputFile in INPUT_FILES.items():
    # read data from file
    freqData = []
    irData = []
    modes = []
    labels = []
    modeFreqs = []
    modeIR = []
    scalingFactors = []
    unscaledFreq = []

    with open(workingDir + moleculeName + ".out", "r") as f:
        for line in f.readlines()[2:]:
            data = line.split("\t")
            freqData.append(float(data[0]))
            irData.append(float(data[1]))
            if len(data) > 3:
                modes.append(int(data[3]))
                labels.append(data[4])
                modeFreqs.append(float(data[5]))
                modeIR.append(float(data[6]))
                scalingFactors.append(float(data[7]))
                unscaledFreq.append(float(data[8]))


    moleculeData.update({
        moleculeName:{
            "freqs":freqData,
            "irData":irData,
            "modes":modes,
            "labels":labels,
            "modeFreqs":modeFreqs,
            "modeIR":modeIR,
            "scalingFactors":scalingFactors,
            "unscaledFreqs":unscaledFreq,
            "peaks":findPeaks([10000 / x for x in freqData], irData, WAV_X_MAX,
                WINDOW_SIZE, N_SIGMA, COALESCE_WINDOW, HIGH_PASS, moleculeName, MPL_PLOT)
        }
    })


    # write to config file with some offset
    sheet = wb["config"]
    sheet["A" + str(configRow)] = moleculeName + " offset"
    sheet["B" + str(configRow)] = currentOffset
    configRow = configRow + 1  # increment current row
    currentOffset += DOFFSET   # increment offset

    # write to data sheet
    wb.create_sheet(moleculeName)
    sheet = wb[moleculeName]

    # Spectrum Data
    sheet["A1"] = "Spectrum"      # write headers
    sheet["A2"] = "Freq (cm^-1)"
    sheet["B2"] = "Wavelength (um)"
    sheet["C2"] = "IR Act"
    sheet["D2"] = "IR Act adj"

    dataRow = 3
    for freq, ir in zip(freqData, irData):
        sheet["A" + str(dataRow)] = freq
        sheet["B" + str(dataRow)] = "=10000/A" + str(dataRow)
        sheet["C" + str(dataRow)] = ir
        sheet["D" + str(dataRow)] = "=C" + str(dataRow) + " + config!B" + str(configRow - 1)  # subtract one to go back

        dataRow += 1

    # Normal Modes
    sheet["F1"] = "Normal Modes"      # write headers
    sheet["F2"] = "Mode"
    sheet["G2"] = "Label"
    sheet["H2"] = "Freq (cm^-1)"
    sheet["I2"] = "IR Act"
    sheet["J2"] = "Scaling Factor"
    sheet["K2"] = "Unscaled freq"

    dataRow = 3
    for mode, label, freq, ir, scale, unscaled in zip(modes, labels, modeFreqs, modeIR, scalingFactors, unscaledFreq):
        sheet["F" + str(dataRow)] = mode
        sheet["G" + str(dataRow)] = label
        sheet["H" + str(dataRow)] = freq
        sheet["I" + str(dataRow)] = ir
        sheet["J" + str(dataRow)] = scale
        sheet["K" + str(dataRow)] = unscaled

        dataRow += 1

    sheet.merge_cells("A1:D1")
    sheet.merge_cells("F1:K1")

    # update sizes
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value



###############################################################################
#
#                         Plots
#
###############################################################################

# label styling
axisFont = CharacterProperties(sz=AXIS_FONT_SIZE * 100)       # multiply by 100, see https://stackoverflow.com/questions/47550555/formatting-chart-data-labels-in-openpyxl
axisTitleFont = CharacterProperties(sz=AXIS_TITLE_FONT_SIZE * 100)  # multiply by 100, see https://stackoverflow.com/questions/47550555/formatting-chart-data-labels-in-openpyxl
labelFont = CharacterProperties(sz=LABEL_FONT_SIZE * 100)     # multiply by 100, see https://stackoverflow.com/questions/47550555/formatting-chart-data-labels-in-openpyxl
pp = ParagraphProperties(defRPr=axisTitleFont)


freqChart = ScatterChart()
freqChart.y_axis.title = "IR Activity"
freqChart.x_axis.title = "Frequency (cm^-1)"
freqChart.x_axis.delete = False
freqChart.y_axis.delete = False
freqChart.x_axis.scaling.max = FREQ_X_MAX
freqChart.x_axis.scaling.min = FREQ_X_MIN
# tick marks
freqChart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axisFont), endParaRPr=axisFont)])
freqChart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axisFont), endParaRPr=axisFont)])
# axis titles
freqChart.x_axis.title.tx.rich.p[0].pPr = pp
freqChart.y_axis.title.tx.rich.p[0].pPr = pp


wavChart = ScatterChart()
wavChart.y_axis.title = "IR Activity"
wavChart.x_axis.title = "Wavelength (µm)"
wavChart.x_axis.delete = False
wavChart.y_axis.delete = False
wavChart.x_axis.scaling.max = WAV_X_MAX
wavChart.x_axis.scaling.min = WAV_X_MIN
# tick marks
wavChart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axisFont), endParaRPr=axisFont)])
wavChart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axisFont), endParaRPr=axisFont)])
# axis titles
wavChart.x_axis.title.tx.rich.p[0].pPr = pp
wavChart.y_axis.title.tx.rich.p[0].pPr = pp


for moleculeName, inputFile in INPUT_FILES.items():
    sheet = wb[moleculeName]
    npoints = len(moleculeData[moleculeName]["freqs"])
    freqXData = Reference(sheet, min_col=1, max_col=1, min_row=3, max_row=2 + npoints)
    wavXData = Reference(sheet, min_col=2, max_col=2, min_row=3, max_row=2 + npoints)

    yData = Reference(sheet, min_col=4, max_col=4, min_row=3, max_row=2 + npoints)

    freqSeries = Series(yData, freqXData, title_from_data=False, title=moleculeName)
    wavSeries = Series(yData, wavXData, title_from_data=False, title=moleculeName)

    # add data labels for peaks
    labels = []
    peaks = moleculeData[moleculeName]["peaks"]
    for i in range(len(wavXData)):
        if i in peaks:
            labels.append(DataLabel(i, showVal=False, showCatName=True, showLeaderLines=True, numFmt="[<0.01]0.E+00;0.00", dLblPos="t"))
        else:
            labels.append(DataLabel(i, showVal=False, showCatName=False, showLeaderLines=True, numFmt="[<0.01]0.E+00;0.00"))


    wavSeries.dLbls = DataLabelList(labels)
    wavSeries.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=labelFont), endParaRPr=labelFont)])

    
    freqChart.series.append(freqSeries)
    wavChart.series.append(wavSeries)


sheet = wb["config"]
sheet.add_chart(freqChart)
sheet.add_chart(wavChart)



###############################################################################
#
#                         Clean up
#
###############################################################################

wb.save(filename=OUTPUT_EXCEL_FILE)

end = timer()

print()
print()
print(f"Finished processing {len(INPUT_FILES.keys())} files in {end - start} seconds")

