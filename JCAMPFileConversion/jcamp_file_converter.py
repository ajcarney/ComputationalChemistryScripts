#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jun 16 21:59:37 2023

@author: aiden
"""

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series


##############################################################################################################
#
#                         Constants
#
##############################################################################################################

# file location information
JCAMP_FILE = "./naphthalene.jdx"
THEORETICAL_DATA_FILE = "IRSpectrum_propene.txt"
OUTPUT_FILE = "test.xlsx"

MOLECULE_NAME = "Propene"  # used to name data sheet


THEORETICAL_Y_OFFSET = 0.3  # offset to add to all 

# headers for data sheet
X_HEADER = "Wavenumber (cm-1)"
Y_HEADER = "IR act"
Y_HEADER2 = "Rescaled IR act"
Y_HEADER3 = "Aid IR act"


# axis titles
X_AXIS = "Frequency (cm^-1)"
Y_AXIS = "IR Activity"




# anything below this line is not to be modified by typical users

##############################################################################################################
#
#                         Parse jcamp File
#
##############################################################################################################

x_data_expr = []
y_data_expr = []

delta_x = -1
x_factor = -1
y_factor = -1
n_points = -1  # sanity checks to make sure everything was read correctly
last_x = -1
first_y = -1


with open(JCAMP_FILE, "r") as f:
    for line in f.readlines():
        if line.startswith("##"):  # reading a comment or some other meta data. try to find the important stuff
            if "XFACTOR" in line:
                x_factor = float(line.split("=")[-1])
            elif "YFACTOR" in line:
                y_factor = float(line.split("=")[-1])
            elif "DELTAX" in line:
                delta_x = float(line.split("=")[-1])
            elif "NPOINTS" in line:
                n_points = int(line.split("=")[-1])
            elif "LASTX" in line:
                last_x = float(line.split("=")[-1])
            elif "FIRSTY" in line:
                first_y = float(line.split("=")[-1])

        elif line[0].isnumeric():  # make sure reading a number
            data = [float(i) for i in line.strip().split(" ")]
            x = data[0]
            for y in data[1:]:
                x_data_expr.append(x)
                y_data_expr.append(y)
                x += delta_x 


# error checking, make sure file was parsed correctly

if any(item == -1 for item in [delta_x, x_factor, y_factor, n_points, last_x, first_y]):
    raise ValueError("Could not read metadata correctly")
if n_points != len(x_data_expr):
    raise ValueError("Did not parse correct number of data points")

x_data_expr = [i * x_factor for i in x_data_expr]
y_data_expr = [i * y_factor for i in y_data_expr]

if last_x != x_data_expr[-1]:
    raise ValueError("Last x value does not match expected from file")

tolerance = 0.01  # y value must match within 1% (tolerance due to floating point representation)
if first_y > y_data_expr[0] * (1 + tolerance) or first_y < y_data_expr[0] * (1 - tolerance):
    raise ValueError("First y value does not match expected from file")




##############################################################################################################
#
#                         Parse theoretical data
# 
##############################################################################################################

x_data_theor = []
y_data_theor = []
max_y_data_theor = 0
rescaled_y_data_theor = []
plot_y_data_theor = []

with open(THEORETICAL_DATA_FILE, "r") as f:
    for line in f.readlines():
        # define first line as the first one that is numeric, this is brittle and sketchy but seems to 
        # be a valid assumption.
        if line[0].isnumeric():
            data = line.split("\t")
            x_data_theor.append(float(data[0]))
            y_data_theor.append(float(data[1]))

max_y_data_theor = max(y_data_theor)
rescaled_y_data_theor = [i * max_y_data_theor for i in y_data_theor]
plot_y_data_theor = [i + THEORETICAL_Y_OFFSET for i in rescaled_y_data_theor]



##############################################################################################################
#
#                         Write data to excel
# 
##############################################################################################################

def excelColToInt(col):
    """
    returns the integer representation of a column in excel by using ascii table offset to calculate value.
    4 index, so A corresponds to 4
    """
    return ord(col) - 64


wb = Workbook()

wb.create_sheet(MOLECULE_NAME)
wb.remove(wb.active)


# write to data sheet
sheet = wb[MOLECULE_NAME]

# write experimental data
exp_row = 2  # start at row 2 for writing data
exp_x_col = "A"
exp_y_col = "B"

sheet[exp_x_col + str(exp_row)] = "Experimental"
exp_row += 1
sheet[exp_x_col + str(exp_row)] = X_HEADER
sheet[exp_y_col + str(exp_row)] = Y_HEADER
exp_row += 1

for x, y in zip(x_data_expr, y_data_expr):
    sheet[exp_x_col + str(exp_row)] = x
    sheet[exp_y_col + str(exp_row)] = y
    exp_row += 1


# write theoretical data
theo_row = 2  # start at row 2 for writing data
theo_x_col = "D"
theo_y_col = "E"
theo_y_col2 = "F"
theo_y_col3 = "G"

scaling_factor_cell_fixed = "$K$6"

sheet[theo_x_col + str(theo_row)] = "Theoretical"
theo_row += 1
sheet[theo_x_col + str(theo_row)] = X_HEADER
sheet[theo_y_col + str(theo_row)] = Y_HEADER
sheet[theo_y_col2 + str(theo_row)] = Y_HEADER2
sheet[theo_y_col3 + str(theo_row)] = Y_HEADER3
theo_row += 1


for x, y in zip(x_data_theor, y_data_theor):
    sheet[theo_x_col + str(theo_row)] = x
    sheet[theo_y_col + str(theo_row)] = y
    sheet[theo_y_col2 + str(theo_row)] = "=" + theo_y_col + str(theo_row) + "*" + scaling_factor_cell_fixed
    sheet[theo_y_col3 + str(theo_row)] = "=" + theo_y_col2 + str(theo_row) + "+" + str(THEORETICAL_Y_OFFSET)

    theo_row += 1


# write the other data. Note this is mostly done on a manual basis so that
# excel formulas can be used rather than everything being performed in python

sheet["I2"] = "Experimental"
sheet["K2"] = "Theoretical"
sheet.merge_cells("I2:J2")
sheet.merge_cells("K2:L2")

sheet["I3"] = "Max " + Y_HEADER
sheet["J3"] = "Max " + X_HEADER
sheet["K3"] = "Max " + Y_HEADER
sheet["L3"] = "Max " + X_HEADER

sheet["I4"] = (
    "=MAX(" + exp_y_col + str(exp_row - len(y_data_expr)) +
    ":" + exp_y_col + str(exp_row - 1) + ")"
)
sheet["J4"] = x_data_expr[y_data_expr.index(max(y_data_expr))]

sheet["K4"] = (
    "=MAX(" + theo_y_col + str(theo_row - len(y_data_theor)) +
    ":" + theo_y_col + str(theo_row - 1) + ")"
)
sheet["L4"] = x_data_theor[y_data_theor.index(max(y_data_theor))]


sheet["J6"] = "Scaling Factor"
sheet["K6"] = "=I4/K4"


sheet["I8"] = "Peak Position Comparison"
sheet.merge_cells("I8:K8")

sheet["I9"] = "Experimental " + X_HEADER
sheet["J9"] = "Theoretical " + X_HEADER
sheet["K9"] = "Difference"




# write plot to sheet

chart1 = ScatterChart()

# Add axis labels
chart1.x_axis.title = X_AXIS
chart1.y_axis.title = Y_AXIS


min_exp_row = exp_row - len(x_data_expr)
max_exp_row = exp_row - 1
exp_x_values = Reference(sheet, 
    min_row=min_exp_row, 
    min_col=excelColToInt(exp_x_col), 
    max_row=max_exp_row, 
    max_col=excelColToInt(exp_x_col)
)
exp_y_values = Reference(sheet, 
    min_row=min_exp_row, 
    min_col=excelColToInt(exp_y_col), 
    max_row=max_exp_row, 
    max_col=excelColToInt(exp_y_col)
)

exp_series = Series(exp_y_values, exp_x_values, title_from_data=False, title="Experimental")


min_theo_row = theo_row - len(y_data_theor)
max_theo_row = theo_row - 1
theo_x_values = Reference(sheet, 
    min_row=min_theo_row, 
    min_col=excelColToInt(theo_x_col), 
    max_row=max_theo_row, 
    max_col=excelColToInt(theo_x_col)
)
theo_y_values = Reference(sheet, 
    min_row=min_theo_row, 
    min_col=excelColToInt(theo_y_col3), 
    max_row=max_theo_row, 
    max_col=excelColToInt(theo_y_col3)
)
theo_series = Series(theo_y_values, theo_x_values, title_from_data=False, title="Theoretical")


chart1.series.append(exp_series)
chart1.append(theo_series)

# style settings
chart1.x_axis.delete = False
chart1.y_axis.delete = False


sheet.add_chart(chart1)



# update sizes
dims = {}
for row in sheet.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) 
for col, value in dims.items():
    sheet.column_dimensions[col].width = value


wb.save(filename=OUTPUT_FILE)



